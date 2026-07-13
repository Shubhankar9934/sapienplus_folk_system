#!/usr/bin/env python3
"""Generate baseline vs final comparison Excel from existing FOLK outputs.

Reads:
  - outputs/countries/*.json  (final deliberated scores)
  - Docs/INDEX OF 171 - WITH FRAMEWORK SCORES.xlsx  (framework baselines + CIs)

Writes:
  - outputs/reports/baseline_vs_final.xlsx

No dependency on the folk package — uses pandas + openpyxl only.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
DIMENSIONS = ("D1", "D2", "D3", "D4")
# Matches src/folk/integrator/engine.py — used to reconstruct pre-clamp placement.
BASELINE_REF_WEIGHT = 0.25
DEFAULT_INFLUENCE = 0.4

DIMENSION_COLUMNS: dict[str, tuple[str, str, str]] = {
    "D1": (
        "factor_1_identity_scaled",
        "factor_1_identity_scaled_ci_lo",
        "factor_1_identity_scaled_ci_hi",
    ),
    "D2": (
        "factor_2_expression_scaled",
        "factor_2_expression_scaled_ci_lo",
        "factor_2_expression_scaled_ci_hi",
    ),
    "D3": (
        "factor_3_structure_scaled",
        "factor_3_structure_scaled_ci_lo",
        "factor_3_structure_scaled_ci_hi",
    ),
    "D4": (
        "factor_4_drive_scaled",
        "factor_4_drive_scaled_ci_lo",
        "factor_4_drive_scaled_ci_hi",
    ),
}

SCORE_COMPARISON_COLUMNS = ["Country", "ISO3"]
for _d in DIMENSIONS:
    SCORE_COMPARISON_COLUMNS.extend([
        f"Baseline_{_d}",
        f"Framework_Low_{_d}",
        f"Framework_High_{_d}",
        f"Recommended_{_d}",
        f"Final_{_d}",
        f"Clamp_Adjustment_{_d}",
        f"Delta_From_Baseline_{_d}",
    ])
SCORE_COMPARISON_COLUMNS.append("Total_Movement")

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def _clean(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load_final_scores(doc: dict) -> dict[str, float] | None:
    """Extract D1-D4 final scores from a country JSON document."""
    methodology = doc.get("methodology") or {}
    final_scores = methodology.get("final_scores") or {}

    out: dict[str, float] = {}
    for dim in DIMENSIONS:
        entry = final_scores.get(dim)
        if isinstance(entry, dict):
            score = _clean(entry.get("score"))
        else:
            score = _clean(entry)
        if score is not None:
            out[dim] = score

    if len(out) == 4:
        return out

    snapshot = (doc.get("cultural_profile") or {}).get("snapshot") or []
    out = {}
    for item in snapshot:
        if not isinstance(item, dict):
            continue
        dim = item.get("dimension")
        score = _clean(item.get("score"))
        if dim in DIMENSIONS and score is not None:
            out[dim] = score

    return out if len(out) == 4 else None


def load_recommended_scores(doc: dict) -> dict[str, float | None]:
    """Integrator pre-CI-clamp recommendation per dimension.

    Prefers the persisted ``range_diagnostics[].integrator_recommendation`` value
    (the exact deterministic integrator placement before _clamp). Only falls back
    to reconstructing it from specialist recommendation + council consensus +
    baseline + evidence credibility when the field is absent (older outputs).
    """
    methodology = doc.get("methodology") or {}
    diagnostics = {
        r.get("dimension"): r
        for r in (methodology.get("range_diagnostics") or [])
        if r.get("dimension") in DIMENSIONS
    }

    # Preferred path (Req 1): read the persisted pre-clamp recommendation.
    persisted: dict[str, float | None] = {}
    have_persisted = False
    for dim in DIMENSIONS:
        val = _clean((diagnostics.get(dim) or {}).get("integrator_recommendation"))
        persisted[dim] = round(val, 1) if val is not None else None
        if val is not None:
            have_persisted = True
    if have_persisted:
        return persisted

    influence = {
        r.get("dimension"): r
        for r in ((methodology.get("specialist") or {}).get("influence_records") or [])
        if r.get("dimension") in DIMENSIONS
    }

    out: dict[str, float | None] = {}
    for dim in DIMENSIONS:
        diag = diagnostics.get(dim)
        if not diag:
            out[dim] = None
            continue

        spec_rec = _clean(diag.get("specialist_recommendation"))
        consensus = _clean(diag.get("council_consensus"))
        baseline = _clean(diag.get("baseline"))

        cred = _clean((influence.get(dim) or {}).get("specialist_influence_weight"))
        if cred is None:
            match = re.search(
                r"evidence credibility ([0-9.]+)",
                str(diag.get("movement_reason") or ""),
            )
            cred = float(match.group(1)) if match else DEFAULT_INFLUENCE

        if spec_rec is not None and consensus is not None:
            evidence_target = cred * spec_rec + (1.0 - cred) * consensus
        elif consensus is not None:
            evidence_target = consensus
        else:
            out[dim] = None
            continue

        if baseline is None:
            provisional = evidence_target
        else:
            w_ref = BASELINE_REF_WEIGHT * (1.0 - cred)
            provisional = (1.0 - w_ref) * evidence_target + w_ref * baseline

        out[dim] = round(provisional, 1)

    return out


def load_framework_baselines(dataset_path: Path) -> dict[str, dict]:
    """Load framework baseline scores and CI bounds keyed by ISO3."""
    df = pd.read_excel(dataset_path)
    by_iso: dict[str, dict] = {}

    for _, row in df.iterrows():
        iso = str(row.get("iso3", "")).strip().upper()
        if not iso or iso == "NAN":
            continue

        baselines: dict[str, float] = {}
        bounds: dict[str, tuple[float, float]] = {}
        complete = True

        for dim in DIMENSIONS:
            score_col, lo_col, hi_col = DIMENSION_COLUMNS[dim]
            baseline = _clean(row.get(score_col))
            lo = _clean(row.get(lo_col))
            hi = _clean(row.get(hi_col))
            if baseline is None:
                complete = False
                break
            baselines[dim] = baseline
            if lo is not None and hi is not None:
                bounds[dim] = (lo, hi)

        if complete:
            by_iso[iso] = {
                "country": str(row.get("country_standard", iso)).strip(),
                "baselines": baselines,
                "bounds": bounds,
            }

    return by_iso


def largest_dimension_change(deltas: dict[str, float]) -> str:
    dim = max(DIMENSIONS, key=lambda d: (abs(deltas[d]), -DIMENSIONS.index(d)))
    delta = deltas[dim]
    sign = "+" if delta >= 0 else "-"
    magnitude = abs(delta)
    text = f"{int(magnitude)}" if magnitude == int(magnitude) else f"{magnitude:.1f}"
    return f"{dim} {sign}{text}"


def total_movement(deltas: dict[str, float]) -> float:
    return sum(abs(deltas[d]) for d in DIMENSIONS)


def load_index_isos(outputs_dir: Path) -> list[str] | None:
    """Read the current run's ISO list from ``outputs/index.json`` (Req 12).

    Reports must reflect ONLY the current run, never orphan JSONs left on disk.
    Returns None when index.json is absent so callers can fall back to a glob.
    """
    index_path = outputs_dir / "index.json"
    if not index_path.is_file():
        return None
    try:
        index = json.loads(index_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    isos = [str(c.get("iso3", "")).strip().upper()
            for c in (index.get("countries") or []) if c.get("iso3")]
    return isos


def build_rows(
    countries_dir: Path,
    framework: dict[str, dict],
    run_isos: list[str] | None = None,
) -> tuple[list[dict], list[str], list[str], list[str]]:
    rows: list[dict] = []
    skipped_incomplete: list[str] = []
    missing_baseline: list[str] = []

    # Prefer the current run's ISO list (index.json); ignore orphan files on disk.
    if run_isos is not None:
        run_set = {i.upper() for i in run_isos}
        paths = [countries_dir / f"{iso}.json" for iso in run_set]
    else:
        paths = sorted(countries_dir.glob("*.json"))
        run_set = {p.stem.upper() for p in paths}

    for path in sorted(paths):
        if not path.is_file():
            skipped_incomplete.append(path.stem.upper())
            continue
        try:
            doc = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            skipped_incomplete.append(path.stem.upper())
            continue

        iso = str(doc.get("iso3") or path.stem).strip().upper()
        country = str(doc.get("country") or iso).strip()
        finals = load_final_scores(doc)

        if finals is None:
            skipped_incomplete.append(iso)
            continue

        fw = framework.get(iso)
        if fw is None or len(fw.get("baselines", {})) != 4:
            missing_baseline.append(iso)
            continue

        baselines = fw["baselines"]
        bounds = fw.get("bounds") or {}
        recommended = load_recommended_scores(doc)
        deltas = {d: finals[d] - baselines[d] for d in DIMENSIONS}

        row: dict = {
            "Country": country,
            "ISO3": iso,
            "Total_Movement": round(total_movement(deltas), 1),
            "_largest_change": largest_dimension_change(deltas),
        }

        for d in DIMENSIONS:
            row[f"Baseline_{d}"] = round(baselines[d], 2)
            b = bounds.get(d)
            row[f"Framework_Low_{d}"] = round(b[0], 2) if b else None
            row[f"Framework_High_{d}"] = round(b[1], 2) if b else None
            rec = recommended.get(d)
            row[f"Recommended_{d}"] = rec
            row[f"Final_{d}"] = int(finals[d])
            # Clamp adjustment = how far the framework CI moved the published score
            # away from the integrator's recommendation (Final - Recommended).
            row[f"Clamp_Adjustment_{d}"] = (round(finals[d] - rec, 1)
                                            if rec is not None else None)
            row[f"Delta_From_Baseline_{d}"] = round(deltas[d], 2)

        rows.append(row)

    rows.sort(key=lambda r: r["Country"].lower())
    missing_outputs = sorted(set(framework) - run_set)
    return rows, skipped_incomplete, missing_baseline, missing_outputs


def _top_n(df: pd.DataFrame, col: str, n: int, ascending: bool) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["Country", "ISO3", col])
    return (
        df.sort_values(col, ascending=ascending)[["Country", "ISO3", col]]
        .head(n)
        .reset_index(drop=True)
    )


def build_summary_rows(df: pd.DataFrame) -> list[list]:
    lines: list[list] = [["Metric", "Value"]]
    lines.append(["Countries Processed", len(df)])

    # Average movement (signed + absolute) per dimension.
    for dim in DIMENSIONS:
        col = f"Delta_From_Baseline_{dim}"
        lines.append([f"Average Movement {dim}", round(df[col].mean(), 2) if len(df) else ""])
        lines.append([
            f"Average Absolute Movement {dim}",
            round(df[col].abs().mean(), 2) if len(df) else "",
        ])

    # Clamp metrics across all dimensions (Req 9 summary).
    clamp_vals: list[float] = []
    clamped = 0
    unclamped = 0
    for dim in DIMENSIONS:
        col = f"Clamp_Adjustment_{dim}"
        if col not in df.columns:
            continue
        for v in df[col].tolist():
            if v is None or (isinstance(v, float) and math.isnan(v)):
                continue
            clamp_vals.append(abs(v))
            if abs(v) >= 0.5:
                clamped += 1
            else:
                unclamped += 1
    avg_clamp = round(sum(clamp_vals) / len(clamp_vals), 2) if clamp_vals else 0
    lines.append(["", ""])
    lines.append(["Average Clamp Adjustment (abs)", avg_clamp])
    lines.append(["Clamped Dimensions", clamped])
    lines.append(["Unclamped Dimensions", unclamped])

    def section(title: str, table: pd.DataFrame, value_col: str) -> None:
        lines.append(["", ""])
        lines.append([title, ""])
        lines.append(["Country", "ISO3", value_col])
        for _, r in table.iterrows():
            val = r[value_col]
            if isinstance(val, float) and val == int(val):
                val = int(val)
            lines.append([r["Country"], r["ISO3"], val])

    section(
        "Top 10 Countries With Largest Total Movement",
        _top_n(df, "Total_Movement", 10, ascending=False),
        "Total_Movement",
    )
    section(
        "Top 10 Countries Closest To Baseline",
        _top_n(df, "Total_Movement", 10, ascending=True),
        "Total_Movement",
    )
    section(
        "Top 10 Countries Furthest From Baseline",
        _top_n(df, "Total_Movement", 10, ascending=False),
        "Total_Movement",
    )

    return lines


def auto_size_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)


def apply_delta_formatting(ws, header_row: list[str]) -> None:
    delta_cols = [i + 1 for i, h in enumerate(header_row) if h.startswith("Delta_")]
    for col_idx in delta_cols:
        col = get_column_letter(col_idx)
        rng = f"{col}2:{col}{ws.max_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["2"], fill=GREEN))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["-2"], fill=RED))
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="between", formula=["-2", "2"], fill=YELLOW),
        )


def format_score_comparison_sheet(ws) -> None:
    header = [cell.value for cell in ws[1]]
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)

    two_decimal_prefixes = ("Baseline_", "Framework_Low_", "Framework_High_")
    one_decimal_prefixes = ("Recommended_", "Clamp_Adjustment_")
    for col_idx, name in enumerate(header, start=1):
        col = get_column_letter(col_idx)
        if name == "Total_Movement":
            for row in range(2, ws.max_row + 1):
                ws[f"{col}{row}"].font = Font(bold=True)
        elif name and name.startswith("Final_"):
            for row in range(2, ws.max_row + 1):
                ws[f"{col}{row}"].number_format = "0"
        elif name and name.startswith(one_decimal_prefixes):
            for row in range(2, ws.max_row + 1):
                ws[f"{col}{row}"].number_format = "0.0"
        elif name and name.startswith(two_decimal_prefixes):
            for row in range(2, ws.max_row + 1):
                ws[f"{col}{row}"].number_format = "0.00"

    apply_delta_formatting(ws, header)
    auto_size_columns(ws)


def format_generic_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
    auto_size_columns(ws)


def write_workbook(
    path: Path,
    comparison_df: pd.DataFrame,
    movement_df: pd.DataFrame,
    summary_rows: list[list],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        comparison_df.to_excel(writer, sheet_name="Score Comparison", index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False, header=False)
        movement_df.to_excel(writer, sheet_name="Movement Ranking", index=False)

        wb = writer.book
        format_score_comparison_sheet(wb["Score Comparison"])
        format_generic_sheet(wb["Movement Ranking"])

        summary_ws = wb["Summary"]
        summary_ws.column_dimensions["A"].width = 45
        summary_ws.column_dimensions["B"].width = 40
        summary_ws.column_dimensions["C"].width = 18


def generate(
    outputs_dir: Path,
    dataset_path: Path,
) -> dict:
    countries_dir = outputs_dir / "countries"
    if not countries_dir.is_dir():
        raise FileNotFoundError(f"Countries directory not found: {countries_dir}")
    if not dataset_path.is_file():
        raise FileNotFoundError(f"Framework spreadsheet not found: {dataset_path}")

    framework = load_framework_baselines(dataset_path)
    # Report only the current run (Req 12): use the ISO list from index.json and
    # ignore orphan JSONs; fall back to a directory glob if index.json is missing.
    run_isos = load_index_isos(outputs_dir)
    rows, skipped, missing_baseline, missing_outputs = build_rows(
        countries_dir, framework, run_isos)

    comparison_df = pd.DataFrame(rows, columns=SCORE_COMPARISON_COLUMNS) if rows else pd.DataFrame(
        columns=SCORE_COMPARISON_COLUMNS
    )

    if rows:
        movement_df = (
            pd.DataFrame(rows)
            .sort_values("Total_Movement", ascending=False)
            .reset_index(drop=True)
        )
        movement_df.insert(0, "Rank", range(1, len(movement_df) + 1))
        movement_df = movement_df[
            ["Rank", "Country", "ISO3", "Total_Movement", "_largest_change"]
        ].rename(columns={"_largest_change": "Largest Dimension Change"})
    else:
        movement_df = pd.DataFrame(
            columns=["Rank", "Country", "ISO3", "Total_Movement", "Largest Dimension Change"]
        )

    summary_rows = build_summary_rows(comparison_df)
    out_path = outputs_dir / "reports" / "baseline_vs_final.xlsx"
    write_workbook(out_path, comparison_df, movement_df, summary_rows)

    clamp_adjustments = 0
    for row in rows:
        for d in DIMENSIONS:
            rec = row.get(f"Recommended_{d}")
            final = row.get(f"Final_{d}")
            if rec is not None and final is not None and abs(final - rec) >= 0.5:
                clamp_adjustments += 1

    return {
        "path": out_path,
        "countries_processed": len(rows),
        "skipped_incomplete": skipped,
        "missing_baseline": missing_baseline,
        "missing_outputs": missing_outputs,
        "clamp_adjustments": clamp_adjustments,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate baseline vs final Excel from FOLK outputs.")
    parser.add_argument("--outputs-dir", type=Path, default=REPO_ROOT / "outputs")
    parser.add_argument(
        "--dataset-path",
        type=Path,
        default=REPO_ROOT / "Docs" / "INDEX OF 171 - WITH FRAMEWORK SCORES.xlsx",
    )
    args = parser.parse_args()

    try:
        result = generate(args.outputs_dir, args.dataset_path)
    except FileNotFoundError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    if result["countries_processed"] == 0:
        print("ERROR: No countries processed — check outputs and spreadsheet.", file=sys.stderr)
        return 1

    print(f"Countries processed:      {result['countries_processed']}")
    print(f"Countries skipped (JSON): {len(result['skipped_incomplete'])}")
    if result["skipped_incomplete"]:
        print(f"  {', '.join(result['skipped_incomplete'])}")
    print(f"Countries missing baseline: {len(result['missing_baseline'])}")
    if result["missing_baseline"]:
        print(f"  {', '.join(result['missing_baseline'])}")
    print(f"Countries missing outputs:  {len(result['missing_outputs'])} (framework rows without JSON)")
    print(f"CI clamp adjustments:       {result['clamp_adjustments']} dimension(s) where Final != Recommended")
    print(f"Workbook:                   {result['path']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
